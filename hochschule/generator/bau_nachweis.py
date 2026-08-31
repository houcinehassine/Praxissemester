# -*- coding: utf-8 -*-
"""Füllt die OTH-Vorlage 'Tätigkeitsnachweis' mit den echten Daten.
Formeln, benannte Bereiche und Layout der Vorlage bleiben unangetastet."""
import os, sys, json, hashlib, re, shutil, zipfile, datetime as dt
HIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HIER)
import openpyxl
from tagestexte import PHASENTEXTE
from hochschultage import START, ENDE, FEIER, KRANK, PRUEFUNGEN, tagtypen

VORLAGE = os.path.join(HIER, "vorlage_taetigkeitsnachweis.xlsx")
ZIEL    = os.path.join(os.path.dirname(HIER), "Hassine_3399727_Tätigkeitsnachweis.xlsx")

# Die Phasen laufen nacheinander und sind über die Anzahl der Arbeitstage
# festgelegt, nicht über feste Datumsgrenzen. Die tatsächlichen Zeiträume
# ergeben sich daraus und werden am Ende ausgegeben.
PHASEN = [
    ("Einarbeitung",          3),
    ("Schraubenlager",       10),
    ("ExcelLagersystem",     20),
    ("CloudAnwendung",        3),
    ("Schweissarbeitsplatz", 14),
    ("Schweisstisch",        14),
    ("Schweisswagen",        13),
    ("Zerspanarbeitsplatz",  15),
    ("Rostschutz",           10),
]

# Die Nettostunden je Tag stehen in stundenplan.py.
from stundenplan import stunden_je_tag

# Vorspann im Tagestext fuer die Tage mit Hochschultermin
VORSPANN = {
    "PP":   "Vorlesung PP 10:00-11:30 an der OTH, davor und danach im Betrieb. ",
    "PPRT": ("Vorlesung PP 10:00-11:30 und Praktikum Regelungstechnik 15:30-17:00 "
             "an der OTH, Vormittag im Betrieb. "),
    "VRT":  "Im Anschluss Praktikum Regelungstechnik 15:30-17:00 an der OTH. ",
}


def medien_zusammenfassen(pfad):
    """Doppelte Bilddateien in der Arbeitsmappe auf eine zusammenziehen.

    Die Vorlage haengt dasselbe Logo an jeden der 13 Druckbloecke. openpyxl
    schreibt es beim Speichern als 13 einzelne Dateien heraus, was die Mappe
    unnoetig aufblaeht. Hier bleibt je Bildinhalt eine Datei uebrig, alle
    Verweise zeigen darauf.
    """
    with zipfile.ZipFile(pfad) as z:
        eintraege = [(i, z.read(i.filename)) for i in z.infolist()]
    erste = {}
    ersetzen = {}
    for info, roh in eintraege:
        if not info.filename.startswith("xl/media/"):
            continue
        schluessel = hashlib.md5(roh).hexdigest()
        if schluessel in erste:
            ersetzen[os.path.basename(info.filename)] = os.path.basename(erste[schluessel])
        else:
            erste[schluessel] = info.filename
    if not ersetzen:
        return 0
    vorlaeufig = pfad + ".tmp"
    with zipfile.ZipFile(vorlaeufig, "w", zipfile.ZIP_DEFLATED) as neu:
        for info, roh in eintraege:
            name = os.path.basename(info.filename)
            if info.filename.startswith("xl/media/") and name in ersetzen:
                continue
            if info.filename.endswith(".rels") or info.filename.endswith("[Content_Types].xml"):
                text = roh.decode("utf-8")
                for alt, behalten in ersetzen.items():
                    text = text.replace("media/" + alt + '"', "media/" + behalten + '"')
                    text = text.replace("/xl/media/" + alt + '"', "/xl/media/" + behalten + '"')
                roh = text.encode("utf-8")
            neu.writestr(info, roh)
    shutil.move(vorlaeufig, pfad)
    return len(ersetzen)


def kopfzeile_reparieren(pfad):
    """Mehrzeilige Kopf-/Fusszeilen aus der Vorlage druckbar machen.

    Die Vorlage kodiert Zeilenumbrueche in Kopf-/Fusszeile als das Excel-
    interne Escape "_x000a_"; openpyxl gibt das beim Speichern unveraendert
    wieder aus. Excel selbst loest das beim Drucken korrekt auf, andere
    Programme (z. B. LibreOffice) zeigen stattdessen den Text "_x000a_"
    woertlich an und schneiden die Fusszeile am unteren Seitenrand ab -
    genau dort, wo die Unterschriftszeile "Datum / Unterschrift und
    Firmenstempel" steht. Deshalb hier durch einen echten Zeilenumbruch
    ersetzt und der untere Seitenrand vergroessert, damit die komplette
    Fusszeile auf jedem Programm sichtbar bleibt, das die Mappe druckt.
    """
    with zipfile.ZipFile(pfad) as z:
        eintraege = [(i, z.read(i.filename)) for i in z.infolist()]
    vorlaeufig = pfad + ".tmp"
    with zipfile.ZipFile(vorlaeufig, "w", zipfile.ZIP_DEFLATED) as neu:
        for info, roh in eintraege:
            if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                text = roh.decode("utf-8")
                if "_x000a_" in text:
                    text = text.replace("_x000a_", "&#10;")
                if "Ausbildungsbetriebes" in text:
                    text = re.sub(r'(<pageMargins\b[^/]*\bbottom=")[^"]*(")', r'\g<1>1.9\g<2>', text)
                roh = text.encode("utf-8")
            neu.writestr(info, roh)
    shutil.move(vorlaeufig, pfad)


def main():
    typen = tagtypen()
    stunden = stunden_je_tag()

    wb = openpyxl.load_workbook(VORLAGE)
    st = wb["Stammdaten"]
    for zelle, wert in {
        "B2": "Herr", "B5": "Hassine", "B8": "Houcine", "B11": "PA6",
        "B14": "3399727", "B17": "houcine1.hassine@hs-regensburg.de",
        "B20": START, "B23": ENDE, "B26": "ja", "B32": 38,
        # Tätigkeitsbereiche aus der Auswahlliste des Blattes "Status"
        "B37": "Logistik und Transport",      # Material- und Schraubenlager
        "B40": "Informatik und Software",     # Lagerbestand-System und Web-Anwendung
        "B43": "Beratung und Planung",        # Schweißarbeitsplatz
        "B46": "Konstruktion",                # Schweißtisch und Schweißmaschinenwagen
        "B49": "Beratung und Planung",        # Zerspanarbeitsplatz
        "I5": "Mechanische Werkstätte Schmidt e.K.", "I8": "Herr",
        "I11": "Halloul", "I14": "Amin",   # Schreibweise nach Vertrag §4
        "I17": "⟨Position im Unternehmen⟩", "I20": "⟨E-Mail Betreuer⟩",
        "I23": "⟨Telefon Betreuer⟩", "I26": "Stiftstraße", "N26": 20,
        "I29": "93343", "L29": "Essing", "I32": "Deutschland",
        "I37": "Anlagen- und Maschinenbau",
    }.items():
        st[zelle] = wert
    for z in ("B20", "B23"):
        st[z].number_format = "DD.MM.YYYY"

    ws = wb["Tätigkeitsnachweis"]
    # Datenzeilen der Vorlage finden (Blöcke à 28 Tage, Kopf wiederholt sich)
    datenzeilen = [r for r in range(16, 600)
                   if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.startswith("=IF(WEEKDAY")]
    # Arbeitstage der Reihe nach auf die Phasen verteilen
    plan = []
    for name, anzahl in PHASEN:
        plan += [name] * anzahl
    rest = {name: list(PHASENTEXTE[name]) for name, _ in PHASEN}
    arbeitstag = 0
    spanne = {}              # Phase -> (erster, letzter Arbeitstag)
    tage = (ENDE - START).days + 1
    protokoll = []

    # In der Vorlage tragen 20 Zeilen in Spalte E ein Zeitformat ('h:mm');
    # eine Stundenzahl würde dort als Uhrzeit erscheinen (5 -> 05:00).
    for r in datenzeilen:
        ws.cell(r, 5).number_format = "0.00"

    for i, r in enumerate(datenzeilen):
        d = START + dt.timedelta(days=i)
        if i >= tage:                       # nach dem letzten Arbeitstag alles leeren
            for sp in (4, 5, 7, 8): ws.cell(r, sp).value = None
            continue
        t = typen[d]
        typ = txt = thema = None; std = None
        if t == "WE":
            pass
        elif t == "F":
            typ, txt = "F", FEIER[d]
        elif t == "K":
            typ, txt = "K", "Krank"
        else:
            # A = nur Betrieb, VA = Betrieb neben Vorlesung, Praktikum oder Pruefung
            typ = "A" if t == "A" else "VA"
            thema = plan[arbeitstag] if arbeitstag < len(plan) else None
            arbeitstag += 1
            if thema:
                erst, _ = spanne.get(thema, (d, d))
                spanne[thema] = (erst, d)
                txt = rest[thema].pop(0) if rest[thema] else None
            if t == "P":
                txt = f"Prüfung {PRUEFUNGEN[d]} an der OTH bis 10:00. " + (txt or "")
            elif t in VORSPANN:
                txt = VORSPANN[t] + (txt or "")
            std = stunden[d]
        ws.cell(r, 4).value = typ
        ws.cell(r, 5).value = std
        ws.cell(r, 7).value = "ja" if t != "A" else "nein"
        ws.cell(r, 8).value = txt
        if typ: protokoll.append((d, typ, std, txt, thema))

    # Verweise und Dokumenteigenschaften der Vorlage bereinigen:
    # Die Vorlage stammt aus einer ausgefüllten Fremddatei; darin hängen an den
    # E-Mail-Feldern noch mailto-Verweise auf die alten Adressen.
    for zelle in ("B17", "I20"):
        st[zelle].hyperlink = None
    st["B17"].hyperlink = "mailto:houcine1.hassine@hs-regensburg.de"
    for eig, wert in (("creator", "Houcine Hassine"), ("lastModifiedBy", "Houcine Hassine"),
                      ("title", "Tätigkeitsnachweis Praxissemester"), ("lastPrinted", None),
                      ("description", None), ("subject", None), ("keywords", None),
                      ("category", None), ("identifier", None), ("language", None)):
        setattr(wb.properties, eig, wert)

    # Excel soll beim Öffnen alle Formeln neu rechnen (openpyxl schreibt keine Ergebniswerte)
    wb.calculation.fullCalcOnLoad = True
    wb.save(ZIEL)
    kopfzeile_reparieren(ZIEL)
    doppelt = medien_zusammenfassen(ZIEL)
    gesamt = sum(p[2] for p in protokoll if p[2])
    wochen = tage / 7
    print(f"geschrieben: {ZIEL}")
    print(f"  Zeilen im Blatt: {len(datenzeilen)} | belegte Tage: {len(protokoll)}"
          f" | doppelte Bilder zusammengefasst: {doppelt}")
    print(f"  Stunden gesamt: {gesamt:.2f} h | Ø {gesamt/wochen:.1f} h/Woche (Stammdaten: 38)")
    from collections import Counter
    print("  Tagestypen:", dict(Counter(p[1] for p in protokoll)))
    leer = [p for p in protokoll if p[1] in ("A", "VA") and not p[3]]
    print("  Arbeitstage ohne Text:", len(leer))
    themen = Counter(p[4] for p in protokoll if p[4])
    print("  Phasen in zeitlicher Reihenfolge:")
    for name, _ in PHASEN:
        a, b = spanne[name]
        print(f"    {name:22} {themen[name]:3} Tage   {a.strftime('%d.%m.')} – {b.strftime('%d.%m.%Y')}")
    json.dump({k: [v[0].isoformat(), v[1].isoformat()] for k, v in spanne.items()},
            open(os.path.join(HIER, "zeitraeume.json"), "w", encoding="utf-8"), indent=1)

if __name__ == "__main__":
    main()
