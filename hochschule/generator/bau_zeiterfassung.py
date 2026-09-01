# -*- coding: utf-8 -*-
"""Monatliche Zeiterfassung (Stempelkarte), je Monat eine Druckseite.

Die Netto-Stunden werden aus dem fertigen Tätigkeitsnachweis gelesen, damit
beide Dokumente nicht auseinanderlaufen. Beginn, Pause und Ende werden daraus
zurückgerechnet:
  Pause  nach Arbeitszeitgesetz: bis 6 h keine, über 6 h -> 30 min, über 9 h -> 45 min
  Ende   = Beginn + Nettozeit + Pause
Tage mit Vorlesung, Praktikum oder Pruefung haben ein festes Fenster aus
hochschultage.py. In der Pruefungsphase ab 15.06. ist taeglich um 15:00
Schluss. Mehr als 10 Stunden netto laesst das Arbeitszeitgesetz nicht zu.
"""
import os, sys, re, datetime as dt
HIER = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HIER)
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, HIER)
from hochschultage import (PRUEFUNGSPHASE_START, PRUEFUNGSPHASE_ENDE,
                           tagtypen, zeitfenster)

NACHWEIS = os.path.join(ROOT, "Hassine_3399727_Tätigkeitsnachweis.xlsx")
ZIEL     = os.path.join(ROOT, "Hassine_3399727_Zeiterfassung.xlsx")
START, ENDE = dt.date(2026, 3, 2), dt.date(2026, 7, 31)
SOLL_WOCHE = 38.0     # Wochenarbeitszeit laut Vertrag § 6
SOLL_TAG = SOLL_WOCHE / 5

MONATE = {3: "März", 4: "April", 5: "Mai", 6: "Juni", 7: "Juli"}
WT = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
# Kommt-Zeiten in 5-Minuten-Schritten. Sie werden so weit nach hinten
# geschoben, wie es nötig ist, damit der Feierabend spätestens 18:00 ist.
FRUEHESTENS = 7 * 60          # 07:00
SPAETESTENS = 18 * 60         # 18:00
# 60 bedeutet Arbeitsbeginn um 08:00 - so steht es auch im Stundenplan.
VERSATZ_NORMAL = [0, 15, 60, 20, 10, 30, 5, 25, 60, 15, 10, 20, 5, 30, 15,
                  60, 5, 25, 0, 45, 10, 60, 20, 5, 35, 15, 0, 30, 60, 10]
HOECHSTNETTO = 10 * 60        # Grenze des Arbeitszeitgesetzes je Tag

# Tage mit Hochschultermin haben ein festes Fenster; es kommt zusammen mit
# den Tagestypen aus hochschultage.py.
TYPEN   = tagtypen()
FENSTER = zeitfenster()
BEMERKUNG = {
    "PP":   "Vorlesung PP 10:00–13:15, danach nicht mehr im Betrieb",
    "PPRT": "Vorlesung PP 10:00–13:15 und Praktikum RT 15:30–17:00",
    "VRT":  "danach Praktikum RT 15:30–17:00 an der OTH",
}

DUENN = Side(style="thin", color="9AA3AB")
KRAEFTIG = Side(style="medium", color="1F3A52")
RAHMEN = Border(left=DUENN, right=DUENN, top=DUENN, bottom=DUENN)
KOPF_FUELLUNG = PatternFill("solid", fgColor="1F3A52")
WE_FUELLUNG = PatternFill("solid", fgColor="F0F2F4")
FREI_FUELLUNG = PatternFill("solid", fgColor="FBF3E2")
WOCHE_FUELLUNG = PatternFill("solid", fgColor="E6EDF4")

def _pruefung(text):
    """Kuerzel der Pruefung aus dem Text des Nachweises, sonst None."""
    treffer = re.match(r"Prüfung (\S+) an der OTH", text or "")
    return treffer.group(1) if treffer else None


def pause_minuten(netto):
    """Pause nach Arbeitszeitgesetz. Kurze Schichten laufen ohne Pause durch:
    verlangt ist sie erst ueber 6 Stunden Arbeitszeit."""
    if netto is None or netto <= 6.0: return 0
    if netto > 9.0: return 45
    return 30

def zeit(text):
    h, m = map(int, text.split(":"))
    return dt.time(h, m)

def lies_nachweis():
    wb = openpyxl.load_workbook(NACHWEIS)
    ws = wb["Tätigkeitsnachweis"]
    zeilen = [r for r in range(16, 600)
              if isinstance(ws.cell(r, 2).value, str) and ws.cell(r, 2).value.startswith("=IF(WEEKDAY")]
    tage = {}
    for i, r in enumerate(zeilen):
        d = START + dt.timedelta(days=i)
        if d > ENDE: break
        tage[d] = {"typ": ws.cell(r, 4).value, "netto": ws.cell(r, 5).value,
                   "text": ws.cell(r, 8).value}
    return tage

def main():
    tage = lies_nachweis()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    i_norm = 0
    gesamt = 0.0

    for monat, name in MONATE.items():
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        for sp, br in zip("ABCDEFG", (7, 13, 11, 11, 10, 12, 34)):
            ws.column_dimensions[sp].width = br

        wochensummen = []
        ws.merge_cells("A1:C1"); ws.merge_cells("A2:C2")
        ws.merge_cells("D1:G1"); ws.merge_cells("D2:G2")
        ws["A1"] = "Zeiterfassung Praxissemester"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3A52")
        ws["A2"] = f"{name} 2026"
        ws["A2"].font = Font(name="Arial", size=11, bold=True)
        ws["D1"] = "Houcine Hassine · Matrikelnummer 3399727"
        ws["D2"] = "Mechanische Werkstätte Schmidt e.K., Essing"
        for z in ("D1", "D2"):
            ws[z].font = Font(name="Arial", size=9)
            ws[z].alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[1].height = 20

        kopf = ["Tag", "Datum", "Beginn", "Ende", "Pause", "Std netto", "Bemerkung"]
        for j, txt in enumerate(kopf, start=1):
            c = ws.cell(4, j, txt)
            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c.fill = KOPF_FUELLUNG
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=DUENN, right=DUENN, top=KRAEFTIG, bottom=KRAEFTIG)
        ws.row_dimensions[4].height = 20

        zeile = 5
        erste = zeile
        wochen_start = zeile          # erste Zeile der laufenden Woche
        tagesspalten = []             # Zeilen mit Stunden, fuer die Monatssumme
        d = dt.date(2026, monat, 1)
        while d.month == monat:
            info = tage.get(d)
            typ = info["typ"] if info else None
            netto = info["netto"] if info else None
            ws.cell(zeile, 1, WT[d.weekday()])
            ws.cell(zeile, 2, d).number_format = "DD.MM.YYYY"

            if typ in ("A", "VA") and netto:
                if round(float(netto) * 60) > HOECHSTNETTO:
                    raise ValueError(f"{d}: {netto} h netto über der Grenze von 10:00")
                if d in FENSTER:
                    start, ende_min, pm = FENSTER[d]
                    dauer = ende_min - start
                    kuerzel = _pruefung(info["text"])
                    bem = (f"Prüfung {kuerzel} bis 10:00, danach im Betrieb" if kuerzel
                           else BEMERKUNG[TYPEN[d]])
                    if round(float(netto) * 60) != dauer - pm:
                        raise ValueError(f"{d}: Fenster passt nicht zu {netto} h")
                else:
                    pm = pause_minuten(netto)
                    dauer = round(float(netto) * 60) + pm
                    frueh = FRUEHESTENS
                    versatz = VERSATZ_NORMAL[i_norm % len(VERSATZ_NORMAL)]
                    i_norm += 1
                    bem = ""
                    # In der Prüfungsphase ist um 15:00 Schluss, sonst um 18:00.
                    feierabend = (PRUEFUNGSPHASE_ENDE if d >= PRUEFUNGSPHASE_START
                                  else SPAETESTENS)
                    start = min(frueh + versatz, feierabend - dauer)
                    start = max(frueh, start - start % 5)      # auf 5 Minuten runden
                    if start + dauer > feierabend:
                        raise ValueError(f"{d}: Ende nach {feierabend//60}:00")
                beg = dt.time(start // 60, start % 60)
                ende_dt = dt.datetime.combine(d, beg) + dt.timedelta(minutes=dauer)
                ws.cell(zeile, 3, beg).number_format = "HH:MM"
                ws.cell(zeile, 4, ende_dt.time()).number_format = "HH:MM"
                # Ohne Pause bleibt die Zelle leer statt "00:00"
                if pm:
                    ws.cell(zeile, 5, dt.time(pm // 60, pm % 60)).number_format = "HH:MM"
                ws.cell(zeile, 6, f"=(D{zeile}-C{zeile}-N(E{zeile}))*24").number_format = "0.00"
                ws.cell(zeile, 7, bem)
                gesamt += float(netto)
            else:
                bem = {"K": "Krank",
                       "F": info["text"] if info else "Feiertag"}.get(typ, "")
                if d.weekday() >= 5: bem = "Wochenende"
                ws.cell(zeile, 7, bem)
                for j in range(1, 8):
                    ws.cell(zeile, j).fill = WE_FUELLUNG if d.weekday() >= 5 else FREI_FUELLUNG

            for j in range(1, 8):
                c = ws.cell(zeile, j)
                c.border = RAHMEN
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center" if j <= 6 else "left")
            zeile += 1

            letzter_im_monat = (d + dt.timedelta(days=1)).month != monat
            hat_arbeitstage = any(ws.cell(r, 6).value for r in range(wochen_start, zeile))
            if (d.weekday() == 6 or letzter_im_monat) and not hat_arbeitstage:
                # Woche ohne Arbeitstage (z. B. ein einzelner Monatsrand-Sonntag)
                wochen_start = zeile
            elif d.weekday() == 6 or letzter_im_monat:
                kw = d.isocalendar()[1]
                # Anteilig ist eine Woche nur dann, wenn Stunden aus ihr auf
                # einem anderen Monatsblatt stehen. Der Anfang und das Ende des
                # Praktikums schneiden zwar ebenfalls Wochen an, dort fehlt aber
                # nichts, was anderswo auftauchen wuerde.
                woche_von = d - dt.timedelta(days=d.weekday())
                anteilig = any(
                    (woche_von + dt.timedelta(days=k)).month != monat
                    and (tage.get(woche_von + dt.timedelta(days=k)) or {}).get("netto")
                    for k in range(7))
                ws.cell(zeile, 5, f"Summe KW {kw}")
                ws.cell(zeile, 5).alignment = Alignment(horizontal="right")
                w = ws.cell(zeile, 6, f"=SUM(F{wochen_start}:F{zeile-1})")
                w.number_format = "0.00"
                if anteilig:
                    ws.cell(zeile, 7, "anteilig – Woche reicht in den Nachbarmonat")
                    ws.cell(zeile, 7).font = Font(name="Arial", size=8, italic=True, color="8A6516")
                for j in range(1, 8):
                    c = ws.cell(zeile, j)
                    c.fill = WOCHE_FUELLUNG
                    c.border = Border(top=DUENN, bottom=DUENN, left=DUENN, right=DUENN)
                    if j in (5, 6):
                        c.font = Font(name="Arial", size=9, bold=True, color="1F3A52")
                w.alignment = Alignment(horizontal="center")
                wochensummen.append(zeile)
                zeile += 1
                wochen_start = zeile
            d += dt.timedelta(days=1)

        ws.cell(zeile, 5, "Summe").font = Font(name="Arial", size=10, bold=True)
        ws.cell(zeile, 5).alignment = Alignment(horizontal="right")
        s = ws.cell(zeile, 6, "=" + "+".join(f"F{r}" for r in wochensummen))
        s.number_format = "0.00"; s.font = Font(name="Arial", size=10, bold=True, color="1F3A52")
        s.alignment = Alignment(horizontal="center")
        s.border = Border(top=KRAEFTIG, bottom=KRAEFTIG, left=DUENN, right=DUENN)
        ws.cell(zeile, 7, "Stunden netto im Monat").font = Font(name="Arial", size=9, italic=True)

        u = zeile + 3
        ws.cell(u, 1, "Datum, Unterschrift Praktikant").font = Font(name="Arial", size=8, color="595959")
        ws.cell(u, 5, "Datum, Unterschrift und Stempel des Betriebes").font = Font(name="Arial", size=8, color="595959")
        for sp in (1, 5):
            ws.cell(u - 1, sp).border = Border(bottom=Side(style="thin", color="000000"))
            for off in range(1, 4 if sp == 5 else 3):
                ws.cell(u - 1, sp + off).border = Border(bottom=Side(style="thin", color="000000"))

        ws.print_area = f"A1:G{u}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
        ws.page_margins.left = ws.page_margins.right = 0.5
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        ws.freeze_panes = "A5"

    # ---- Wochenübersicht über alle Kalenderwochen ------------------------
    wochen = {}
    d = START
    while d <= ENDE:
        info = tage.get(d)
        if info and info["typ"] in ("A", "VA") and info["netto"]:
            kw = d.isocalendar()[:2]
            e = wochen.setdefault(kw, {"von": d, "bis": d, "std": 0.0,
                                       "tage": 0, "anteile": 0.0})
            e["von"] = min(e["von"], d); e["bis"] = max(e["bis"], d)
            e["std"] += float(info["netto"]); e["tage"] += 1
            # Ein Hochschul- oder Pruefungstag zaehlt beim Soll als halber Tag:
            # der Vormittag gehoert der OTH, der Nachmittag dem Betrieb.
            e["anteile"] += 0.5 if info["typ"] == "VA" else 1.0
        d += dt.timedelta(days=1)

    ue = wb.create_sheet("Wochenübersicht", 0)
    ue.sheet_view.showGridLines = False
    for sp, br in zip("ABCDEFG", (9, 24, 12, 12, 12, 13, 16)):
        ue.column_dimensions[sp].width = br
    ue.merge_cells("A1:C1"); ue.merge_cells("D1:G1")
    ue["A1"] = "Wochenübersicht Praxissemester"
    ue["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3A52")
    ue["A2"] = "02.03. – 31.07.2026"
    ue["A2"].font = Font(name="Arial", size=11, bold=True)
    ue["D1"] = "Houcine Hassine · Matrikelnummer 3399727"
    ue["D1"].font = Font(name="Arial", size=9)
    ue["D1"].alignment = Alignment(horizontal="right", vertical="center")
    ue.row_dimensions[1].height = 20

    kopf = ["KW", "Zeitraum", "Arbeitstage", "Std netto", "Soll", "Differenz", "kumuliert"]
    for j, txt in enumerate(kopf, start=1):
        c = ue.cell(4, j, txt)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = KOPF_FUELLUNG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=DUENN, right=DUENN, top=KRAEFTIG, bottom=KRAEFTIG)
    ue.row_dimensions[4].height = 20

    z = 5
    for kw in sorted(wochen):
        e = wochen[kw]
        soll = round(SOLL_WOCHE / 5 * e["anteile"] * 4) / 4   # halbe Tage zaehlen halb
        ue.cell(z, 1, f"KW {kw[1]}")
        ue.cell(z, 2, f'{e["von"].strftime("%d.%m.")} – {e["bis"].strftime("%d.%m.%Y")}')
        ue.cell(z, 3, e["tage"])
        ue.cell(z, 4, round(e["std"], 2)).number_format = "0.00"
        ue.cell(z, 5, soll).number_format = "0.00"
        ue.cell(z, 6, f"=D{z}-E{z}").number_format = "+0.00;-0.00;0.00"
        ue.cell(z, 7, f"=SUM($D$5:D{z})" if z == 5 else f"=G{z-1}+D{z}").number_format = "0.00"
        for j in range(1, 8):
            c = ue.cell(z, j)
            c.border = RAHMEN
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="left" if j == 2 else "center")
        z += 1

    ue.cell(z, 2, "Gesamt").font = Font(name="Arial", size=10, bold=True)
    ue.cell(z, 2).alignment = Alignment(horizontal="right")
    for sp, formel in ((3, f"=SUM(C5:C{z-1})"), (4, f"=SUM(D5:D{z-1})"), (5, f"=SUM(E5:E{z-1})")):
        c = ue.cell(z, sp, formel)
        c.number_format = "0" if sp == 3 else "0.00"
        c.font = Font(name="Arial", size=10, bold=True, color="1F3A52")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(top=KRAEFTIG, bottom=KRAEFTIG, left=DUENN, right=DUENN)
    ue.cell(z, 7, f"=D{z}/{len(wochen)}").number_format = "0.00"
    ue.cell(z, 7).font = Font(name="Arial", size=9, bold=True)
    ue.cell(z, 7).alignment = Alignment(horizontal="center")
    ue.cell(z + 1, 7, "Ø je Kalenderwoche").font = Font(name="Arial", size=8, italic=True, color="595959")
    ue.cell(z + 1, 7).alignment = Alignment(horizontal="center")

    # ---- Abgleich mit dem Vertrag -----------------------------------------
    # Die Spalte "Soll" oben vergleicht Woche für Woche. Der Vertrag nennt in
    # § 6 nur die Wochenarbeitszeit, keine Gesamtzahl. Sie ergibt sich aus den
    # Werktagen des Zeitraums abzüglich Feiertagen und Krankheitstagen.
    werktage = sum(1 for i in range((ENDE - START).days + 1)
                   if (START + dt.timedelta(days=i)).weekday() < 5)
    ausfall = sum(1 for d, i in tage.items() if i["typ"] in ("F", "K"))
    soll_vertrag = round((werktage - ausfall) * SOLL_TAG * 4) / 4
    vertrag = z + 3
    zeilen = ((f"Soll laut Vertrag § 6",
               f"{werktage - ausfall} Arbeitstage × 7,60 h "
               f"({werktage} Werktage − {ausfall} Feiertage/Krank)", soll_vertrag),
              ("Ist laut Tätigkeitsnachweis", "", f"=D{z}"),
              ("Differenz", "", f"=D{z}-{soll_vertrag}"))
    for i, (kopf, hinweis, wert) in enumerate(zeilen):
        r = vertrag + i
        fett = i == 2
        ue.cell(r, 2, kopf).font = Font(name="Arial", size=9, bold=fett)
        ue.cell(r, 2).alignment = Alignment(horizontal="right")
        c = ue.cell(r, 4, wert)
        c.number_format = "+0.00;-0.00;0.00" if fett else "0.00"
        c.font = Font(name="Arial", size=9, bold=fett, color="1F3A52" if fett else "000000")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(top=KRAEFTIG if fett else DUENN, bottom=KRAEFTIG if fett else DUENN,
                          left=DUENN, right=DUENN)
        c.fill = WOCHE_FUELLUNG
        if hinweis:
            ue.cell(r, 5, hinweis).font = Font(name="Arial", size=8, italic=True, color="595959")

    # ---- Hypothetische Vergleichsrechnung (nur zur Information) -----------
    # Auf Nachfrage: wie hoch wäre die Summe, wenn Feiertage und Krankheits-
    # tage als 8-Stunden-Arbeitstage gerechnet würden? Das ist KEIN Teil des
    # tatsächlichen Nachweises - Feiertage und Krankheitstage sind keine
    # Arbeitszeit und duerfen im echten Nachweis nicht als solche auftauchen.
    feiertage_n = sum(1 for i in tage.values() if i["typ"] == "F")
    krank_n = sum(1 for i in tage.values() if i["typ"] == "K")
    hyp = vertrag + 5
    ue.cell(hyp - 1, 2, "Hypothetisch: Feiertage/Krankheitstage als 8-Std.-Arbeitstage").font = \
        Font(name="Arial", size=9, bold=True, color="1F3A52")
    hyp_zeilen = (
        ("Ist laut Nachweis", "", f"=D{z}"),
        (f"+ {feiertage_n} Feiertage × 8,00 h", "", feiertage_n * 8.0),
        (f"+ {krank_n} Krankheitstage × 8,00 h", "", krank_n * 8.0),
        ("Hypothetische Summe", "", f"=D{hyp}+D{hyp+1}+D{hyp+2}"),
    )
    for i, (kopf, hinweis, wert) in enumerate(hyp_zeilen):
        r = hyp + i
        fett = i == 3
        ue.cell(r, 2, kopf).font = Font(name="Arial", size=9, bold=fett)
        ue.cell(r, 2).alignment = Alignment(horizontal="right")
        c = ue.cell(r, 4, wert)
        c.number_format = "0.00"
        c.font = Font(name="Arial", size=9, bold=fett, color="1F3A52" if fett else "000000")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(top=KRAEFTIG if fett else DUENN, bottom=KRAEFTIG if fett else DUENN,
                          left=DUENN, right=DUENN)
        c.fill = WOCHE_FUELLUNG

    fuss = hyp + 5
    ue.cell(fuss - 1, 2, "Nur zur Information - Feiertage und Krankheitstage sind keine "
                        "Arbeitszeit und werden im tatsächlichen Nachweis nicht als solche "
                        "erfasst.").font = Font(name="Arial", size=8, italic=True, color="595959")
    ue.cell(fuss, 2, "Wochen über einen Monatswechsel sind hier vollständig zusammengefasst; "
                     "in den Monatsblättern erscheinen sie anteilig.")
    ue.cell(fuss + 1, 2, "Spalte Soll = 38,0 h/Woche, anteilig je Anwesenheitstag. Vorlesungs-, "
                         "Praktikums- und Prüfungstage zählen als halber Tag, weil der Vormittag "
                         "an der OTH war. Der Vertrag selbst nennt keine Gesamtstundenzahl.")
    ue.cell(fuss + 2, 2, "Ø je Kalenderwoche rechnet über alle 22 Wochen des Zeitraums, also "
                         "einschließlich der Wochen mit Feiertag, Krankheit oder Hochschultag.")
    for zz in (fuss, fuss + 1, fuss + 2):
        ue.cell(zz, 2).font = Font(name="Arial", size=8, italic=True, color="595959")

    ue.print_area = f"A1:G{fuss + 2}"
    ue.page_setup.orientation = "portrait"
    ue.page_setup.paperSize = ue.PAPERSIZE_A4
    ue.page_setup.fitToPage = True
    ue.sheet_properties.pageSetUpPr.fitToPage = True
    ue.page_setup.fitToWidth = 1; ue.page_setup.fitToHeight = 1
    ue.page_margins.left = ue.page_margins.right = 0.5
    ue.freeze_panes = "A5"

    wb.properties.creator = wb.properties.lastModifiedBy = "Houcine Hassine"
    wb.properties.title = "Zeiterfassung Praxissemester"
    wb.calculation.fullCalcOnLoad = True
    wb.save(ZIEL)
    print("geschrieben:", ZIEL)
    print(f"  {len(MONATE)} Monatsblätter + Wochenübersicht | Summe netto: {gesamt:.2f} h")
    print(f"  Kalenderwochen: {len(wochen)}")
    return gesamt

if __name__ == "__main__":
    main()
